import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font


def get_paths():
    base_path = os.path.dirname(os.path.dirname(__file__))
    data_path = os.path.join(base_path, "data", "data.csv")
    output_excel = os.path.join(base_path, "output", "reporte.xlsx")
    chart_path = os.path.join(base_path, "output", "duracion_por_equipo.png")

    return data_path, output_excel, chart_path


def load_data(path):
    return pd.read_csv(path)


def process_data(df):
    total = len(df)
    equipos = df["equipment"].nunique()

    mtto = df["technician"].value_counts().sort_values(ascending=False)
    duracion = df.groupby("equipment")["duration_hours"].mean().sort_values(ascending=False)
    estado = df["status"].value_counts()

    return total, equipos, mtto, duracion, estado


def create_chart(duracion, chart_path):
    duracion.plot(kind="bar")
    plt.title("Duración promedio por equipo")
    plt.ylabel("Horas")
    plt.xticks(rotation=0)
    plt.tight_layout()
    plt.savefig(chart_path)
    plt.close()


def generate_excel(output_path, total, equipos, df, estado, mtto, duracion):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        resumen = pd.DataFrame({
            "Indicador": [
                "Total mantenimientos",
                "Equipos únicos",
                "Duración promedio (hrs)",
                "Mantenimientos completados",
                "Mantenimientos retrasados"
            ],
            "Valor": [
                total,
                equipos,
                round(df["duration_hours"].mean(), 2),
                estado.get("Completed", 0),
                estado.get("Delayed", 0)
            ]
        })

        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        mtto.to_frame(name="Cantidad").to_excel(writer, sheet_name="Por Ingeniero")
        duracion.to_frame(name="Horas promedio").to_excel(writer, sheet_name="Duración por equipo")
        estado.to_frame(name="Cantidad").to_excel(writer, sheet_name="Estado")


def format_excel(output_path, chart_path):
    wb = load_workbook(output_path)

    # Insertar gráfica
    ws = wb["Duración por equipo"]
    img = Image(chart_path)
    ws.add_image(img, "E2")

    # Ajustar columnas
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max(max_length + 2, 15)

        # Estilo encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Centrar datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

    wb.save(output_path)


def main():
    data_path, output_path, chart_path = get_paths()

    df = load_data(data_path)

    total, equipos, mtto, duracion, estado = process_data(df)

    create_chart(duracion, chart_path)

    generate_excel(output_path, total, equipos, df, estado, mtto, duracion)

    format_excel(output_path, chart_path)


if __name__ == "__main__":
    main()